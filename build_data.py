#!/usr/bin/env python3
"""Exporta mad-map-data-v2.xlsx -> mad-map-data.json para las superficies HTML.

Implementa Modelo IV (embedding estructural + proyección PCA):
- Cada sublínea recibe un vector de características construido desde sus
  relaciones estructurales (matriz de afinidad declarada + área + línea-madre
  + investigadores que la cultivan).
- PCA via SVD reduce ese vector a posiciones 2D y 3D normalizadas en [-1, 1].
- Las líneas se posicionan en el centroide de sus sublíneas.
- Los investigadores en el centroide de las sublíneas que cultivan.
- Las posiciones quedan baked en el JSON; el frontend sólo escala y renderiza.

Fuente del embedding (Source C en el spec elicit): graph-based, sin
dependencias externas. La fuente A (texto) y D (híbrida) son evolución
posterior — basta sumar un vector textual al feature stack.

Estructura del JSON:
- sello: foco + texto del sello formativo elegido
- entities: lineas, sublineas, areas, modos, salidas, laboratorios,
  investigadores, temas (cada nodo lleva ahora `position.2d` y `position.3d`)
- relations: las tablas m:n del modelo
- edges: pre-computadas, agrupadas por tipo (a..g del spec)
"""

import json
from pathlib import Path
from collections import defaultdict

import numpy as np
from openpyxl import load_workbook

ROOT = Path(__file__).parent
XLSX = ROOT / "mad-map-data-v2.xlsx"
OUT = ROOT / "mad-map-data.json"


def read_table(wb, sheet_name):
    """Lee una tabla autodetectando la fila de encabezados.

    Las hojas creadas por build_xlsx.py ponen los encabezados en fila 2
    (cuando hay sólo título) o en fila 4 (cuando hay título + nota + fila
    en blanco). Buscamos la primera fila donde todas las columnas tienen
    valor; esa es la fila de encabezados.
    """
    ws = wb[sheet_name]
    header_row = None
    for r in range(1, 8):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is not None for v in vals) and len(vals) > 1:
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"No se detectó fila de encabezados en {sheet_name}")
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {h: ws.cell(r, c).value for c, h in enumerate(headers, 1) if h}
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows


def compute_positions(sublineas, lineas, investigadores, areas, proximidad, sub_to_invs):
    """Modelo IV: posiciones de cada nodo desde embedding estructural.

    Estrategia:
      1. Construir matriz de características por sublínea (afinidad declarada
         + área + línea + investigadores que la cultivan).
      2. PCA via SVD numérica → dos primeras componentes (2D) y tres (3D).
      3. Normalizar a [-1, 1].
      4. Líneas = centroide de sus sublíneas.
      5. Investigadores = centroide de las sublíneas que cultivan.

    Anota in-place `position = {"2d": [x, y], "3d": [x, y, z]}` en cada nodo.
    """
    n = len(sublineas)
    if n == 0:
        return
    sub_idx = {s["id"]: i for i, s in enumerate(sublineas)}

    # 1.a Matriz de afinidad declarada (Sublínea ↔ Sublínea)
    A = np.zeros((n, n), dtype=float)
    for p in proximidad:
        ai = sub_idx.get(p.get("sublinea_a_id"))
        bi = sub_idx.get(p.get("sublinea_b_id"))
        af = p.get("afinidad")
        if ai is None or bi is None or af is None:
            continue
        A[ai, bi] = float(af)
        A[bi, ai] = float(af)
    np.fill_diagonal(A, 1.0)

    # 1.b Área one-hot
    area_ids = [a["id"] for a in areas]
    area_oh = np.zeros((n, len(area_ids)))
    for i, s in enumerate(sublineas):
        if s["area"] in area_ids:
            area_oh[i, area_ids.index(s["area"])] = 1.0

    # 1.c Línea-madre one-hot
    linea_ids = [l["id"] for l in lineas]
    linea_oh = np.zeros((n, len(linea_ids)))
    for i, s in enumerate(sublineas):
        if s["linea"] in linea_ids:
            linea_oh[i, linea_ids.index(s["linea"])] = 1.0

    # 1.d Investigadores que cultivan cada sublínea
    inv_ids = [inv["id"] for inv in investigadores]
    inv_oh = np.zeros((n, len(inv_ids)))
    for i, s in enumerate(sublineas):
        for inv_id in sub_to_invs.get(s["id"], set()):
            if inv_id in inv_ids:
                inv_oh[i, inv_ids.index(inv_id)] = 1.0

    # 2. Concatenar features con pesos
    #    - afinidad declarada: peso 1.5 (señal semántica directa, curada)
    #    - línea-madre: peso 2.0 (cluster fuerte por pertenencia troncal)
    #    - área: peso 1.0 (clasificación institucional ortogonal)
    #    - investigadores: peso 0.5 (señal de proximidad por cuerpo académico)
    F = np.hstack([
        A * 1.5,
        linea_oh * 2.0,
        area_oh * 1.0,
        inv_oh * 0.5,
    ])

    # 3. PCA via SVD
    F_centered = F - F.mean(axis=0, keepdims=True)
    U, S, _ = np.linalg.svd(F_centered, full_matrices=False)
    components = U * S  # (n × min(n, d))

    sub_pos_2d = components[:, :2].copy()
    sub_pos_3d = components[:, :3].copy() if components.shape[1] >= 3 else \
                 np.hstack([components[:, :2], np.zeros((n, 1))])

    # 4. Normalizar cada eje a [-1, 1] (usando max abs por dimensión)
    def normalize(coords):
        max_abs = np.max(np.abs(coords), axis=0)
        max_abs = np.where(max_abs > 1e-9, max_abs, 1.0)
        return coords / max_abs

    sub_pos_2d = normalize(sub_pos_2d)
    sub_pos_3d = normalize(sub_pos_3d)

    for i, s in enumerate(sublineas):
        s["position"] = {
            "2d": sub_pos_2d[i].tolist(),
            "3d": sub_pos_3d[i].tolist(),
        }

    # 5. Líneas: centroide de sus sublíneas
    sublines_by_line = defaultdict(list)
    for i, s in enumerate(sublineas):
        if s.get("linea"):
            sublines_by_line[s["linea"]].append(i)

    for l in lineas:
        indices = sublines_by_line.get(l["id"], [])
        if indices:
            l["position"] = {
                "2d": sub_pos_2d[indices].mean(axis=0).tolist(),
                "3d": sub_pos_3d[indices].mean(axis=0).tolist(),
            }
        else:
            l["position"] = {"2d": [0.0, 0.0], "3d": [0.0, 0.0, 0.0]}

    # 6. Investigadores: centroide de las sublíneas que cultivan
    inv_to_subs_idx = defaultdict(list)
    for i, s in enumerate(sublineas):
        for inv_id in sub_to_invs.get(s["id"], set()):
            inv_to_subs_idx[inv_id].append(i)

    for inv in investigadores:
        indices = inv_to_subs_idx.get(inv["id"], [])
        if indices:
            inv["position"] = {
                "2d": sub_pos_2d[indices].mean(axis=0).tolist(),
                "3d": sub_pos_3d[indices].mean(axis=0).tolist(),
            }
        else:
            # Investigador sin sublíneas mapeadas: posición neutral, fácil de identificar
            inv["position"] = {"2d": [0.0, 0.0], "3d": [0.0, 0.0, 0.0]}


def main():
    wb = load_workbook(XLSX, data_only=True)

    # ---- Entidades ----
    lineas = [{
        "id": r["id"],
        "nombre": r["nombre"],
        "descripcion": r["descripción"],
        "estado": r.get("estado", ""),
    } for r in read_table(wb, "01_Lineas")]

    sublineas = [{
        "id": r["id"],
        "nombre": r["nombre"],
        "linea": r["linea_id"],
        "area": r["area_id"],
        "notas": r.get("notas") or "",
    } for r in read_table(wb, "02_Sublineas")]

    areas = [{
        "id": r["código"],
        "nombre": r["nombre"],
        "descripcion": r["descripción"],
    } for r in read_table(wb, "03_Areas")]

    modos = [{
        "id": r["id"],
        "nombre": r["nombre"],
        "descripcion": r["descripción"],
    } for r in read_table(wb, "04_Modos")]

    salidas = [{
        "id": r["id"],
        "nombre": r["nombre"],
        "descripcion": r["descripción"],
    } for r in read_table(wb, "05_Salidas")]

    laboratorios = [{
        "id": r["id"],
        "nombre": r["nombre"],
        "descripcion": r["descripción"],
    } for r in read_table(wb, "06_Laboratorios")]

    investigadores = [{
        "id": r["id"],
        "nombre": r["nombre"],
        "area_principal": r["área_principal"],
        "perfil_url": r["perfil_casiopea"],
    } for r in read_table(wb, "07_Investigadores")]

    temas = [{
        "id": r["id"],
        "investigador": r["investigador_id"],
        "texto": r["tema"],
    } for r in read_table(wb, "08_Temas")]

    # ---- Relaciones m:n ----
    sublinea_tema = read_table(wb, "09_Sublinea_Tema")
    lab_linea = read_table(wb, "10_Lab_Linea")
    lab_salida = read_table(wb, "11_Lab_Salida")
    inv_lab = read_table(wb, "12_Investigador_Lab")
    inv_modo = read_table(wb, "13_Investigador_Modo")
    linea_modo = read_table(wb, "14_Linea_Modo")

    # Proximidad: descartar pares marcados como DESCARTADO
    proximidad_raw = read_table(wb, "18_Proximidad_Tematica")
    proximidad = [p for p in proximidad_raw if p.get("estado") != "DESCARTADO"]

    # ---- Sello formativo (variante elegida) ----
    sello_ws = wb["17_Sello"]
    sello = {"foco": "", "texto": ""}
    for r in range(5, sello_ws.max_row + 1):
        foco = sello_ws.cell(r, 2).value
        if foco and "ELEGIDO" in str(foco):
            sello = {
                "foco": str(foco).replace(" (ELEGIDO)", ""),
                "texto": sello_ws.cell(r, 3).value,
            }
            break

    # ---- Pre-computar aristas (a..g del spec) ----
    edges = {
        "jerarquica": [],           # (a) Sublinea -> Linea
        "coautoria": [],            # (b) Investigador -> Sublinea (vía Tema)
        "coinvestigacion": [],      # (c) Sublinea <-> Sublinea (perfiles compartidos)
        "sosten_lab": [],           # (d) Lab -> Linea
        "afinidad_lab": [],         # (e) Sublinea <-> Sublinea (mismo Lab)
        "coincidencia_modo": [],    # (f) Sublinea <-> Sublinea (mismo Modo predominante)
        "proximidad_semantica": [], # (g) Sublinea <-> Sublinea (declarada)
    }

    # (a) Jerárquica: cada Sublínea tiene una Línea madre
    for s in sublineas:
        if s["linea"]:
            edges["jerarquica"].append({
                "source": s["id"],
                "target": s["linea"],
            })

    # (b) Coautoría: Investigador -> Sublínea vía Temas
    tema_to_inv = {t["id"]: t["investigador"] for t in temas}
    inv_to_sublines = defaultdict(set)
    for st in sublinea_tema:
        sub_id = st.get("sublinea_id")
        tema_id = st.get("tema_id")
        if not sub_id or not tema_id:
            continue
        inv_id = tema_to_inv.get(tema_id)
        if inv_id:
            inv_to_sublines[inv_id].add(sub_id)
    for inv_id, subs in inv_to_sublines.items():
        for sub_id in subs:
            edges["coautoria"].append({"source": inv_id, "target": sub_id})

    # (c) Coinvestigación: pares de Sublíneas que comparten Investigador
    sub_to_invs = defaultdict(set)
    for inv_id, subs in inv_to_sublines.items():
        for sub_id in subs:
            sub_to_invs[sub_id].add(inv_id)
    sub_ids_sorted = sorted(sub_to_invs.keys())
    for i, a in enumerate(sub_ids_sorted):
        for b in sub_ids_sorted[i + 1:]:
            shared = sub_to_invs[a] & sub_to_invs[b]
            if shared:
                edges["coinvestigacion"].append({
                    "source": a, "target": b, "weight": len(shared),
                })

    # (d) Sostén Lab: Laboratorio -> Línea
    for ll in lab_linea:
        edges["sosten_lab"].append({
            "source": ll["lab_id"],
            "target": ll["linea_id"],
        })

    # (e) Afinidad por Lab: Sublíneas cuya Línea-madre comparte Lab
    linea_to_labs = defaultdict(set)
    for ll in lab_linea:
        linea_to_labs[ll["linea_id"]].add(ll["lab_id"])
    linea_to_sublines = defaultdict(set)
    for s in sublineas:
        if s["linea"]:
            linea_to_sublines[s["linea"]].add(s["id"])
    afinidad_lab_pairs = set()
    # dos sublíneas comparten lab si sus líneas-madre tienen al menos un lab en común
    sub_to_labs = defaultdict(set)
    for s_id, _ in [(s["id"], s["linea"]) for s in sublineas]:
        pass
    for s in sublineas:
        if s["linea"] and s["linea"] in linea_to_labs:
            sub_to_labs[s["id"]] = linea_to_labs[s["linea"]]
    sub_ids_with_lab = sorted([s for s in sub_to_labs if sub_to_labs[s]])
    for i, a in enumerate(sub_ids_with_lab):
        for b in sub_ids_with_lab[i + 1:]:
            if sub_to_labs[a] & sub_to_labs[b]:
                afinidad_lab_pairs.add((a, b))
    for a, b in afinidad_lab_pairs:
        edges["afinidad_lab"].append({"source": a, "target": b})

    # (f) Coincidencia de modo: Sublíneas cuya Línea-madre comparte Modo predominante
    linea_to_modos = defaultdict(set)
    for lm in linea_modo:
        if lm.get("nivel") == "predominante":
            linea_to_modos[lm["linea_id"]].add(lm["modo_id"])
    sub_to_modos = defaultdict(set)
    for s in sublineas:
        if s["linea"]:
            sub_to_modos[s["id"]] = linea_to_modos.get(s["linea"], set())
    modo_pairs = set()
    sub_ids_with_modo = sorted([s for s in sub_to_modos if sub_to_modos[s]])
    for i, a in enumerate(sub_ids_with_modo):
        for b in sub_ids_with_modo[i + 1:]:
            if sub_to_modos[a] & sub_to_modos[b]:
                modo_pairs.add((a, b))
    for a, b in modo_pairs:
        edges["coincidencia_modo"].append({"source": a, "target": b})

    # (g) Proximidad semántica declarada
    for p in proximidad:
        edges["proximidad_semantica"].append({
            "source": p["sublinea_a_id"],
            "target": p["sublinea_b_id"],
            "weight": p["afinidad"],
        })

    # ---- Modelo IV: posiciones desde embedding estructural ----
    # `sub_to_invs` ya se construyó arriba para la arista "coinvestigacion"
    compute_positions(
        sublineas=sublineas,
        lineas=lineas,
        investigadores=investigadores,
        areas=areas,
        proximidad=proximidad,
        sub_to_invs=sub_to_invs,
    )

    # ---- Salida ----
    data = {
        "version": "modelo-iv-source-c",
        "sello": sello,
        "lineas": lineas,
        "sublineas": sublineas,
        "areas": areas,
        "modos": modos,
        "salidas": salidas,
        "laboratorios": laboratorios,
        "investigadores": investigadores,
        "temas": temas,
        "relations": {
            "sublinea_tema": sublinea_tema,
            "lab_linea": lab_linea,
            "lab_salida": lab_salida,
            "inv_lab": inv_lab,
            "inv_modo": inv_modo,
            "linea_modo": linea_modo,
            "proximidad": proximidad,
        },
        "edges": edges,
    }

    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2))
    print(f"OK: {OUT}")
    print(f"  Líneas: {len(lineas)}")
    print(f"  Sublíneas: {len(sublineas)}")
    print(f"  Investigadores: {len(investigadores)}")
    print(f"  Temas: {len(temas)}")
    print(f"  Áreas: {len(areas)} | Modos: {len(modos)} | Salidas: {len(salidas)} | Labs: {len(laboratorios)}")
    print(f"  Edges:")
    for k, v in edges.items():
        print(f"    {k}: {len(v)}")


if __name__ == "__main__":
    main()
