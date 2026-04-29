"""Genera la planilla relacional de Líneas/Sublíneas/Áreas/Modos/Salidas/Labs/Investigadores/Temas
para el nuevo régimen del MAD-map. Estructura: 1 hoja por entidad + hojas de relación m:n,
más sello formativo y matriz de proximidad temática.

Re-ejecutar este script REGENERA la planilla desde cero. Los ajustes manuales hechos en
las hojas (p. ej. cambios de afinidad o estado en 18_Proximidad_Tematica) se PIERDEN.
Para que los cambios sobrevivan, reflejarlos en PROXIMIDAD_PARES o SELLO_VARIANTES de
este mismo archivo.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = str(Path(__file__).parent / "mad-map-data-v2.xlsx")

# ---------------------------------------------------------------------------
# Datos de referencia
# ---------------------------------------------------------------------------

AREAS = [
    ("ECH", "Extensión, Ciudad y Habitabilidad",
     "Marco amplio: ciudad, territorio, urbanización, infraestructura urbana, patrimonio territorial, borde costero, prácticas urbanas colectivas."),
    ("EAA", "Educación, Espacio y Aprendizaje",
     "Marco amplio: espacios educativos, pedagogía del diseño, arquitectura como medio didáctico, contextos vulnerables, historia y teoría como campo formativo, archivo y memoria."),
    ("FCT", "Forma, Cultura y Tecnología",
     "Marco amplio: diseño de interacción, servicios y UX, accesibilidad e inclusión, accesibilidad cognitiva, diseño social y territorial, oficio y artesanía, sistemas inteligentes, transferencia tecnológica."),
]

MODOS = [
    ("MOD-01", "Historiografía",
     "Investigación basada en fuentes, archivo, biografía intelectual, genealogías de obra y acervo."),
    ("MOD-02", "Teoría crítica",
     "Producción conceptual, ensayo argumentativo, lectura crítica de prácticas y discursos."),
    ("MOD-03", "Investigación proyectual",
     "Investigación a través del proyecto: el proyecto-obra como dispositivo de generación de conocimiento."),
]

SALIDAS = [
    ("SAL-IND", "Industria",
     "Investigación aplicada, innovación y desarrollo tecnológico; transferencia al sector productivo."),
    ("SAL-ACA", "Academia",
     "Formación e investigación disciplinar, innovación docente, generación de conocimiento disciplinar."),
    ("SAL-EST", "Estado",
     "Asuntos y políticas públicas; vínculo con instituciones del Estado."),
]

LABORATORIOS = [
    ("LAB-NAI", "Núcleo de Accesibilidad e Inclusión",
     "Accesibilidad cognitiva, codiseño, comunicación aumentativa/alternativa, vida independiente, participación inclusiva."),
    ("LAB-AFL", "Aconcagua Fablab",
     "Fabricación digital, modelado paramétrico, transferencia tecnológica distribuida, sistemas inteligentes, hacer-crear."),
    ("LAB-PYT", "Personas y territorios",
     "Ciudad, territorio, urbanización, patrimonio territorial y comunidades; intersección persona–lugar."),
    ("LAB-PMD", "Patrimonio moderno",
     "Historia y crítica de la arquitectura moderna, acervo de la e[ad] y Ciudad Abierta, archivo y memoria material."),
    ("LAB-UAF", "Urbanismo afectivo",
     "Deriva, prácticas urbanas colectivas, cooperativismo, reuso del patrimonio en abandono, investigación-acción."),
]

# Líneas troncales — borrador propuesto a partir de la hipótesis de compactación.
# El usuario debe decidir si son 3 o 4. Se proponen 4, marcadas como BORRADOR.
LINEAS = [
    ("LIN-01", "Personas, interacción y sistemas inclusivos",
     "Investigación sobre las condiciones de quien habita: accesibilidad e inclusión, codiseño, comunicación aumentativa y alternativa, diseño de interacción y servicios, sistemas inteligentes orientados a la diversidad funcional, la autonomía y la democracia.",
     "consolidada"),
    ("LIN-02", "Habitar, infraestructura y ecologías territoriales",
     "Investigación sobre las condiciones de dónde se habita: ciudad, urbanización, ecología política, patrimonio territorial, borde costero, infraestructura urbana, vulnerabilidad y adaptación, prácticas urbanas colectivas y reuso del territorio.",
     "consolidada"),
    ("LIN-03", "Teoría e historia de la arquitectura y el diseño",
     "Investigación sobre las condiciones desde dónde se piensa: acervo histórico de la e[ad] y Ciudad Abierta, historia y crítica de la arquitectura moderna, teoría del proyecto, conceptos disciplinares como hospitalidad y vacío arquitectónico.",
     "consolidada"),
    ("LIN-04", "Oficio, tecnologías y aprendizaje proyectual",
     "Investigación sobre las condiciones con qué se hace y cómo se transmite: métodos de diseño, oficio, saberes técnicos, fabricación digital y transferencia tecnológica, espacios educativos y pedagogía del proyecto.",
     "consolidada"),
]

# ---------------------------------------------------------------------------
# Investigadores (perfiles de la EAD-PUCV declarados en Casiopea/ANID)
# ---------------------------------------------------------------------------
# (id, nombre, area_principal, perfil_url, estado_perfil)
INVESTIGADORES = [
    ("INV-01", "Ursula Exss", "EAA", "https://wiki.ead.pucv.cl/Ursula_Exss", "consolidado Casiopea + ANID"),
    ("INV-02", "Katherine Exss", "FCT", "https://wiki.ead.pucv.cl/Katherine_Exss", "consolidado Casiopea + ANID"),
    ("INV-03", "Jorge Ferrada", "ECH", "https://wiki.ead.pucv.cl/Jorge_Ferrada", "bio + cursos + proyectos + tesis"),
    ("INV-04", "Andrés Garcés", "ECH", "https://wiki.ead.pucv.cl/Andr%C3%A9s_Garc%C3%A9s", "línea explícita"),
    ("INV-05", "Iván Ivelic", "ECH", "https://wiki.ead.pucv.cl/Iv%C3%A1n_Ivelic", "línea explícita + lectura detallada"),
    ("INV-06", "David Luza", "ECH", "https://wiki.ead.pucv.cl/David_Luza", "consolidado Casiopea + ANID"),
    ("INV-07", "Adriana Marín", "ECH", "https://wiki.ead.pucv.cl/Adriana_Mar%C3%ADn", "líneas explícitas"),
    ("INV-08", "Álvaro Mercado", "ECH", "https://wiki.ead.pucv.cl/%C3%81lvaro_Mercado", "líneas explícitas"),
    ("INV-09", "Jaime Reyes", "EAA", "https://wiki.ead.pucv.cl/Jaime_Reyes", "tema/área explícita"),
    ("INV-10", "Rodrigo Saavedra", "EAA", "https://wiki.ead.pucv.cl/Rodrigo_Saavedra", "líneas explícitas"),
    ("INV-11", "Daniela Salgado", "FCT", "https://wiki.ead.pucv.cl/Daniela_Salgado", "líneas explícitas"),
    ("INV-12", "Herbert Spencer", "FCT", "https://wiki.ead.pucv.cl/Herbert_Spencer", "líneas explícitas"),
    ("INV-13", "Alfred Thiers", "ECH", "https://wiki.ead.pucv.cl/Alfred_Thiers", "líneas explícitas"),
    ("INV-14", "Lorena Herrera", "ECH", "https://wiki.ead.pucv.cl/Lorena_Herrera", "ANID"),
    ("INV-15", "Juan Carlos Jeldes", "FCT", "https://wiki.ead.pucv.cl/Juan_Carlos_Jeldes", "lectura detallada de perfil"),
    ("INV-16", "Michèle Wilkomirsky", "EAA", "https://wiki.ead.pucv.cl/Mich%C3%A8le_Wilkomirsky", "ANID"),
    ("INV-17", "Óscar Andrade", "EAA", "https://wiki.ead.pucv.cl/%C3%93scar_Andrade", "ANID"),
    ("INV-18", "Anna Braghini", "EAA", "https://wiki.ead.pucv.cl/Anna_Braghini", "deducción razonable"),
    ("INV-19", "Arturo Chicano", "FCT", "https://wiki.ead.pucv.cl/Arturo_Chicano", "lectura detallada de perfil"),
    ("INV-20", "Sylvia Arriagada", "FCT", "https://wiki.ead.pucv.cl/Sylvia_Arriagada", "línea explícita + lectura detallada"),
    ("INV-21", "Marcelo Araya", "FCT", "https://wiki.ead.pucv.cl/Marcelo_Araya", "deducción fundada"),
    ("INV-22", "Emanuela Di Felice", "ECH", "https://wiki.ead.pucv.cl/Emanuela_Di_Felice", "consolidado por lectura detallada"),
]

# ---------------------------------------------------------------------------
# Sublíneas (banco depurado normalizado)
# Cada sublínea pertenece a EXACTAMENTE una Línea troncal (por decisión del usuario).
# El campo linea_id es BORRADOR.
# ---------------------------------------------------------------------------
SUBLINEAS = [
    # (id, nombre, linea_id_borrador, area_id, notas)
    ("SUB-01", "Accesibilidad cognitiva y codiseño",                         "LIN-01", "FCT", ""),
    ("SUB-02", "Accesibilidad e inclusión",                                  "LIN-01", "FCT", ""),
    ("SUB-03", "Acervo histórico de Ciudad Abierta y de la Escuela de Arquitectura y Diseño", "LIN-03", "EAA", ""),
    ("SUB-04", "Adaptabilidad urbana ante riesgos costeros",                 "LIN-02", "ECH", ""),
    ("SUB-05", "Adaptación formal ante desastres",                           "LIN-02", "ECH", ""),
    ("SUB-06", "Arquitectura como medio didáctico",                          "LIN-04", "EAA", "migró a LIN-04 (aprendizaje proyectual)"),
    ("SUB-07", "Comunicación aumentativa y alternativa",                     "LIN-01", "FCT", ""),
    ("SUB-08", "Diseño arquitectónico",                                      "LIN-03", "EAA", ""),
    ("SUB-09", "Diseño de espacios educativos",                              "LIN-04", "EAA", "migró a LIN-04"),
    ("SUB-10", "Diseño de interacción, servicios y experiencia de usuario",  "LIN-01", "FCT", ""),
    ("SUB-11", "Diseño y vida independiente",                                "LIN-01", "FCT", ""),
    ("SUB-12", "El sentido de la hospitalidad",                              "LIN-03", "EAA", ""),
    ("SUB-13", "El vacío arquitectónico",                                    "LIN-03", "EAA", ""),
    ("SUB-14", "Espacios de aprendizaje en contextos vulnerables",           "LIN-04", "EAA", "migró a LIN-04"),
    ("SUB-15", "Espacios de estimulación temprana",                          "LIN-04", "EAA", "migró a LIN-04"),
    ("SUB-16", "Espacios educativos en contextos vulnerables",               "LIN-04", "EAA", "migró a LIN-04"),
    ("SUB-17", "Formación en pensamiento y acción creativa",                 "LIN-04", "FCT", ""),
    ("SUB-18", "Industrias creativas e innovación social para el desarrollo local", "LIN-04", "FCT", ""),
    ("SUB-19", "Medios de aprendizaje en arquitectura y diseño",             "LIN-04", "EAA", "migró a LIN-04"),
    ("SUB-20", "Métodos de diseño",                                          "LIN-04", "FCT", ""),
    ("SUB-21", "Perspectivas decoloniales",                                  "LIN-02", "FCT", "podría leerse en LIN-01 según énfasis"),
    ("SUB-22", "Prácticas colectivas",                                       "LIN-02", "EAA", "transversal — colocar donde tenga más respaldo"),
    ("SUB-23", "Programas educativos y espacios escolares",                  "LIN-04", "EAA", "migró a LIN-04"),
    ("SUB-24", "Proyecto de arquitectura escolar",                           "LIN-04", "EAA", "migró a LIN-04"),
    ("SUB-25", "Saberes técnicos análogos",                                  "LIN-04", "FCT", ""),
    ("SUB-26", "Teoría e historia de la arquitectura",                       "LIN-03", "EAA", ""),
    ("SUB-27", "Transferencia de medios tecnológicos con aplicación al emprendimiento local", "LIN-04", "FCT", ""),
    ("SUB-28", "Travesías y geopoética",                                     "LIN-02", "ECH", ""),
    ("SUB-29", "Urbanización, urbanismo y ecología política",                "LIN-02", "ECH", ""),
    ("SUB-30", "Vulnerabilidad física de asentamientos costeros",            "LIN-02", "ECH", ""),
    # ----- Nuevas sublíneas (SUB-31..SUB-52) -----
    ("SUB-31", "Vivienda, financiarización y políticas habitacionales",       "LIN-02", "ECH", "nueva"),
    ("SUB-32", "Comunes, comunalidad y resiliencias socioecológicas",         "LIN-02", "ECH", "nueva"),
    ("SUB-33", "Movilidad, infraestructura urbana y equipamiento",            "LIN-02", "ECH", "nueva"),
    ("SUB-34", "Ciudad-teatro y apropiación escénica del espacio público",    "LIN-02", "ECH", "nueva"),
    ("SUB-35", "Naturaleza, paisaje y patrimonio natural",                    "LIN-02", "ECH", "nueva"),
    ("SUB-36", "Patrimonio arquitectónico y rehabilitación",                  "LIN-02", "ECH", "nueva"),
    ("SUB-37", "Urbanismo afectivo, deriva y reuso del patrimonio",           "LIN-02", "ECH", "nueva"),
    ("SUB-38", "Investigación-acción y artes urbanas",                        "LIN-02", "ECH", "nueva"),
    ("SUB-39", "Evaluación social y políticas públicas de inversión",         "LIN-02", "ECH", "nueva"),
    ("SUB-40", "Mobiliario, vida cotidiana y materialidad de la obra",        "LIN-04", "FCT", "nueva"),
    ("SUB-41", "Fabricación digital, modelado paramétrico y fablabs",         "LIN-04", "FCT", "nueva — núcleo Jeldes"),
    ("SUB-42", "Diseño para la democracia y comunicación ciudadana",          "LIN-01", "FCT", "nueva"),
    ("SUB-43", "Comunicación visual, diseño editorial y exposición material", "LIN-04", "FCT", "nueva"),
    ("SUB-44", "Diseño social, territorial y patrimonio cultural",            "LIN-02", "FCT", "nueva"),
    ("SUB-45", "Confort, bienestar y habitabilidad personal",                 "LIN-02", "FCT", "nueva"),
    ("SUB-46", "Máquinas expresivas, algoritmos y arte tecnológico",          "LIN-04", "FCT", "nueva"),
    ("SUB-47", "Tecnología, sociedad y techné",                               "LIN-03", "FCT", "nueva"),
    ("SUB-48", "Arte y arquitectura",                                         "LIN-03", "EAA", "nueva"),
    ("SUB-49", "Arquitectura moderna y habitar latinoamericano",              "LIN-03", "EAA", "nueva"),
    ("SUB-50", "Poesía y oficio",                                             "LIN-03", "EAA", "nueva"),
    ("SUB-51", "Fundamentos poéticos y artísticos del diseño",                "LIN-03", "EAA", "nueva"),
    ("SUB-52", "Reforma escolar y enseñanza de la arquitectura",              "LIN-04", "EAA", "nueva"),
    ("SUB-53", "Sistemas inteligentes e IA en diseño",                        "LIN-01", "FCT", "nueva — separa IA aplicada al diseño democrático/inclusivo de fabricación digital"),
]

# ---------------------------------------------------------------------------
# Temas declarados por cada investigador (texto crudo desde Casiopea/ANID)
# Estructura: (investigador_id, tema)
# ---------------------------------------------------------------------------
TEMAS_RAW = [
    # Ursula Exss
    ("INV-01", "Espacios educativos"),
    ("INV-01", "Arquitectura escolar"),
    ("INV-01", "Relación entre arquitectura y aprendizaje"),
    ("INV-01", "Reforma escolar de 1965 en Chile"),
    ("INV-01", "Teoría de la arquitectura"),
    ("INV-01", "Diseño arquitectónico"),
    ("INV-01", "Diseño de espacios educativos"),
    # Katherine Exss
    ("INV-02", "Diseño de interacción y experiencia de usuarios"),
    ("INV-02", "Metodologías y métodos de diseño"),
    ("INV-02", "Accesibilidad e inclusión"),
    ("INV-02", "Confort térmico y bienestar"),
    ("INV-02", "Accesibilidad cognitiva"),
    ("INV-02", "Investigación inclusiva"),
    # Jorge Ferrada
    ("INV-03", "Patrimonio arquitectónico y territorial"),
    ("INV-03", "Borde costero y ciudad-puerto"),
    ("INV-03", "Movilidad y orden urbano en Valparaíso"),
    ("INV-03", "Rehabilitación patrimonial"),
    ("INV-03", "Muros de contención como estructura urbana"),
    ("INV-03", "Frentes marítimos y ribereños"),
    ("INV-03", "Territorio marítimo y portuario"),
    # Andrés Garcés
    ("INV-04", "Ciudad-teatro"),
    ("INV-04", "Cultura y ciudad"),
    ("INV-04", "Fenómenos sociales informales"),
    ("INV-04", "Espacio y forma arquitectónica"),
    ("INV-04", "Apropiación del espacio público como acto escénico"),
    # Iván Ivelic
    ("INV-05", "Áreas naturales protegidas"),
    ("INV-05", "Relación artificio y naturaleza"),
    ("INV-05", "Equipamientos de interpretación, educación y servicios"),
    ("INV-05", "Áreas urbanas"),
    ("INV-05", "Equipamiento cultural"),
    ("INV-05", "Patrimonio natural y cultural"),
    ("INV-05", "Sustentabilidad del patrimonio"),
    ("INV-05", "Arquitectura como interfaz entre habitante y naturaleza"),
    ("INV-05", "Creación colectiva y experimentación arquitectónica"),
    ("INV-05", "Proyecto y obra como generación de conocimiento"),
    # David Luza
    ("INV-06", "Habitabilidad en infraestructura urbana"),
    ("INV-06", "Experimentación arquitectónica en Ciudad Abierta"),
    ("INV-06", "Temas urbanos"),
    ("INV-06", "Equipamiento e infraestructura urbana"),
    ("INV-06", "Ciudad y territorio"),
    ("INV-06", "Movilidad sustentable"),
    ("INV-06", "Habitabilidad en infraestructura urbana técnica y natural"),
    # Adriana Marín
    ("INV-07", "Vivienda en arriendo"),
    ("INV-07", "Financiarización de la vivienda"),
    ("INV-07", "Teoría urbana"),
    ("INV-07", "Metodologías de investigación cualitativa"),
    # Álvaro Mercado
    ("INV-08", "Ecología política y diseño de la transición"),
    ("INV-08", "Urbanización capitalista y extractiva"),
    ("INV-08", "Comunes y comunalidad"),
    ("INV-08", "Decolonialidad y urbanización"),
    ("INV-08", "Frontera, relacionalidad y diferencia"),
    ("INV-08", "Transformaciones socioecológicas"),
    ("INV-08", "Resiliencias humanas y más-que-humanas"),
    # Jaime Reyes
    ("INV-09", "Relación poesía-oficio"),
    ("INV-09", "Seminario de América"),
    ("INV-09", "Archivo histórico"),
    ("INV-09", "Investigación en posgrado dentro de Educación, Espacio y Aprendizaje"),
    # Rodrigo Saavedra
    ("INV-10", "Arquitectura como medio didáctico"),
    ("INV-10", "Capacidad didáctica de la obra"),
    ("INV-10", "Espacios de aprendizaje en contextos socialmente vulnerables"),
    ("INV-10", "Estimulación temprana"),
    # Daniela Salgado
    ("INV-11", "Diseño y artesanía"),
    ("INV-11", "Diseño social y territorial"),
    ("INV-11", "Diseño decolonial"),
    ("INV-11", "Interacción entre oficio y cultura"),
    ("INV-11", "Objetos, contexto social y cultural"),
    ("INV-11", "Patrimonio cultural material e inmaterial"),
    # Herbert Spencer
    ("INV-12", "Diseño universal desde la accesibilidad cognitiva"),
    ("INV-12", "Comunicación visual"),
    ("INV-12", "Participación inclusiva"),
    ("INV-12", "Diseño de sistemas inteligentes"),
    ("INV-12", "IA como material de diseño"),
    ("INV-12", "Diseño para la democracia"),
    ("INV-12", "Plataformas de comunicación ciudadana"),
    ("INV-12", "Comunicación aumentativa alternativa"),
    ("INV-12", "Diseño y vida independiente"),
    ("INV-12", "Diseño de interacción y servicios"),
    # Alfred Thiers
    ("INV-13", "Habitabilidad"),
    ("INV-13", "Mobiliario y lugar"),
    ("INV-13", "Actos de la vida urbana y cotidiana"),
    ("INV-13", "Materialidad y obra"),
    ("INV-13", "Sistemas de proyecto desde origen a implementación"),
    # Lorena Herrera
    ("INV-14", "Inversión pública y evaluación social de proyectos"),
    ("INV-14", "Desarrollo de metodologías para la evaluación social de proyectos"),
    ("INV-14", "Valoración de beneficios sociales generados por proyectos de inversión"),
    # Juan Carlos Jeldes
    ("INV-15", "Fabricación digital"),
    ("INV-15", "Fablabs"),
    ("INV-15", "Modelado paramétrico"),
    ("INV-15", "Diseño tecnológico aplicado"),
    ("INV-15", "Tecnología y sociedad"),
    ("INV-15", "Aprendizaje desde el hacer-crear"),
    ("INV-15", "Investigación proyectual"),
    ("INV-15", "Mediación territorial y hospitalidad en lo habitable"),
    ("INV-15", "Modelado paramétrico interpretativo"),
    ("INV-15", "Transferencia tecnológica distribuida"),
    # Michèle Wilkomirsky
    ("INV-16", "Pedagogía del Diseño"),
    ("INV-16", "Diseño de información visual"),
    # Óscar Andrade
    ("INV-17", "Diseño arquitectónico"),
    ("INV-17", "Educación en arquitectura"),
    ("INV-17", "Teoría de la arquitectura"),
    ("INV-17", "Arte y arquitectura"),
    # Anna Braghini
    ("INV-18", "Historia y crítica de la arquitectura moderna"),
    ("INV-18", "Arquitectura moderna latinoamericana"),
    ("INV-18", "Estudio del concepto de espacio"),
    ("INV-18", "Habitar moderno"),
    ("INV-18", "Arquitectura del Cono Sur"),
    ("INV-18", "Patrimonio cultural de cuadernos-bitácora de la e[ad]"),
    # Arturo Chicano
    ("INV-19", "Máquinas expresivas"),
    ("INV-19", "Algoritmos cinéticos"),
    ("INV-19", "Techné y mundo"),
    ("INV-19", "Epistemología del diseño"),
    ("INV-19", "Innovación social y tecnológica"),
    ("INV-19", "Transferencias disciplinares y formativas"),
    ("INV-19", "Observación como transmisión de ethos"),
    ("INV-19", "Habitabilidad y biodiversidad"),
    # Sylvia Arriagada
    ("INV-20", "Fundamentos poéticos y artísticos del diseño"),
    ("INV-20", "Diseño gráfico editorial"),
    ("INV-20", "Montajes expositivos"),
    ("INV-20", "Heredad creativa del diseño en Travesías"),
    ("INV-20", "Acervo fundacional de la Escuela"),
    ("INV-20", "Poética del oficio"),
    ("INV-20", "Actos poéticos"),
    ("INV-20", "Contenidos patrimoniales y su exposición material"),
    # Marcelo Araya
    ("INV-21", "Interacción física en espacios expositivos"),
    ("INV-21", "Interacción como material de diseño"),
    ("INV-21", "Taller topológico multiescalar"),
    ("INV-21", "Formas resistentes del diseño"),
    ("INV-21", "Estructuras cinéticas y mecánico-expresivas"),
    ("INV-21", "Visualización territorial"),
    ("INV-21", "Patrimonio urbano y memoria material"),
    ("INV-21", "Multiterritorialidad y frontera"),
    # Emanuela Di Felice
    ("INV-22", "Deriva urbana"),
    ("INV-22", "Caminar como práctica estética y artística"),
    ("INV-22", "Artes urbanas"),
    ("INV-22", "Investigación-acción"),
    ("INV-22", "Urbanismo afectivo"),
    ("INV-22", "Movimientos autogestionados"),
    ("INV-22", "Reutilización del patrimonio abandonado"),
    ("INV-22", "Cooperativismo"),
    ("INV-22", "Reuso del patrimonio público en abandono"),
]

TEMAS = [(f"TEM-{i+1:03d}", inv, t) for i, (inv, t) in enumerate(TEMAS_RAW)]

# ---------------------------------------------------------------------------
# Investigador ↔ Lab (m:n) — primera lectura desde el material
# ---------------------------------------------------------------------------
INV_LAB = [
    ("INV-01", "LAB-PMD"),  # Ursula Exss — historia/teoría arquitectura
    ("INV-02", "LAB-NAI"),  # Katherine Exss — accesibilidad cognitiva
    ("INV-02", "LAB-AFL"),  # Katherine Exss — interacción/metodologías
    ("INV-03", "LAB-PYT"),  # Jorge Ferrada — patrimonio territorial
    ("INV-03", "LAB-PMD"),  # Jorge Ferrada — patrimonio
    ("INV-04", "LAB-UAF"),  # Andrés Garcés — apropiación urbana
    ("INV-04", "LAB-PYT"),  # Andrés Garcés
    ("INV-05", "LAB-PYT"),  # Iván Ivelic — territorio/patrimonio natural
    ("INV-06", "LAB-PYT"),  # David Luza — ciudad y territorio
    ("INV-06", "LAB-UAF"),  # David Luza — habitabilidad urbana
    ("INV-07", "LAB-PYT"),  # Adriana Marín — vivienda y teoría urbana
    ("INV-08", "LAB-UAF"),  # Álvaro Mercado — ecología política
    ("INV-08", "LAB-PYT"),  # Álvaro Mercado
    ("INV-09", "LAB-PMD"),  # Jaime Reyes — archivo, Seminario de América
    ("INV-10", "LAB-NAI"),  # Rodrigo Saavedra — contextos vulnerables, estimulación temprana
    ("INV-11", "LAB-NAI"),  # Daniela Salgado — diseño social, decolonial
    ("INV-11", "LAB-PYT"),  # Daniela Salgado — territorio
    ("INV-12", "LAB-NAI"),  # Herbert Spencer — accesibilidad cognitiva
    ("INV-12", "LAB-AFL"),  # Herbert Spencer — sistemas inteligentes
    ("INV-13", "LAB-UAF"),  # Alfred Thiers — habitabilidad, mobiliario, vida cotidiana
    ("INV-14", "LAB-PYT"),  # Lorena Herrera — evaluación social proyectos
    ("INV-15", "LAB-AFL"),  # Juan Carlos Jeldes — fablab, fab digital
    ("INV-16", "LAB-PMD"),  # Michèle Wilkomirsky — pedagogía del diseño
    ("INV-17", "LAB-PMD"),  # Óscar Andrade — historia/teoría
    ("INV-18", "LAB-PMD"),  # Anna Braghini — moderna latinoamericana
    ("INV-19", "LAB-AFL"),  # Arturo Chicano — máquinas, algoritmos
    ("INV-20", "LAB-PMD"),  # Sylvia Arriagada — acervo fundacional
    ("INV-21", "LAB-AFL"),  # Marcelo Araya — interacción, estructuras cinéticas
    ("INV-21", "LAB-PMD"),  # Marcelo Araya — patrimonio urbano y memoria
    ("INV-22", "LAB-UAF"),  # Emanuela Di Felice — urbanismo afectivo (homónimo)
]

# ---------------------------------------------------------------------------
# Investigador ↔ Modo (m:n) — primera lectura
# ---------------------------------------------------------------------------
INV_MODO = [
    ("INV-01", "MOD-02"), ("INV-01", "MOD-01"),
    ("INV-02", "MOD-03"), ("INV-02", "MOD-02"),
    ("INV-03", "MOD-01"), ("INV-03", "MOD-02"),
    ("INV-04", "MOD-02"), ("INV-04", "MOD-03"),
    ("INV-05", "MOD-03"), ("INV-05", "MOD-02"),
    ("INV-06", "MOD-03"),
    ("INV-07", "MOD-02"),
    ("INV-08", "MOD-02"),
    ("INV-09", "MOD-01"), ("INV-09", "MOD-02"),
    ("INV-10", "MOD-03"), ("INV-10", "MOD-02"),
    ("INV-11", "MOD-03"), ("INV-11", "MOD-02"),
    ("INV-12", "MOD-03"),
    ("INV-13", "MOD-03"),
    ("INV-14", "MOD-02"),
    ("INV-15", "MOD-03"),
    ("INV-16", "MOD-02"), ("INV-16", "MOD-03"),
    ("INV-17", "MOD-02"), ("INV-17", "MOD-01"),
    ("INV-18", "MOD-01"), ("INV-18", "MOD-02"),
    ("INV-19", "MOD-03"), ("INV-19", "MOD-02"),
    ("INV-20", "MOD-01"), ("INV-20", "MOD-03"),
    ("INV-21", "MOD-03"), ("INV-21", "MOD-02"),
    ("INV-22", "MOD-03"), ("INV-22", "MOD-02"),
]

# ---------------------------------------------------------------------------
# Lab ↔ Línea (m:n) — primera lectura
# ---------------------------------------------------------------------------
LAB_LINEA = [
    ("LAB-NAI", "LIN-01"),
    ("LAB-AFL", "LIN-01"),
    ("LAB-AFL", "LIN-04"),
    ("LAB-PYT", "LIN-02"),
    ("LAB-UAF", "LIN-02"),
    ("LAB-PMD", "LIN-03"),
]

# ---------------------------------------------------------------------------
# Lab ↔ Salida (m:n) — los labs operacionalizan las salidas de transferencia
# (declaración del usuario; primera lectura para revisión)
# ---------------------------------------------------------------------------
LAB_SALIDA = [
    ("LAB-NAI", "SAL-EST"),
    ("LAB-NAI", "SAL-ACA"),
    ("LAB-AFL", "SAL-IND"),
    ("LAB-AFL", "SAL-ACA"),
    ("LAB-PYT", "SAL-EST"),
    ("LAB-PYT", "SAL-ACA"),
    ("LAB-PMD", "SAL-ACA"),
    ("LAB-UAF", "SAL-EST"),
    ("LAB-UAF", "SAL-ACA"),
]

# ---------------------------------------------------------------------------
# Sublínea ↔ Tema (m:n) — un tema puede colgar de varias sublíneas
# Mapeo de primera lectura, alta confianza solamente. Lo demás queda para curaduría.
# ---------------------------------------------------------------------------
def find_tema(tema_text, inv_id):
    for tid, iid, t in TEMAS:
        if iid == inv_id and t == tema_text:
            return tid
    return None

SUB_TEMA_MAPPINGS = [
    # (subid, [(inv_id, tema_text), ...])
    ("SUB-01", [
        ("INV-02", "Accesibilidad cognitiva"),
        ("INV-12", "Diseño universal desde la accesibilidad cognitiva"),
    ]),
    ("SUB-02", [
        ("INV-02", "Accesibilidad e inclusión"),
        ("INV-02", "Investigación inclusiva"),
        ("INV-12", "Participación inclusiva"),
    ]),
    ("SUB-03", [
        ("INV-09", "Archivo histórico"),
        ("INV-18", "Patrimonio cultural de cuadernos-bitácora de la e[ad]"),
        ("INV-20", "Acervo fundacional de la Escuela"),
        ("INV-20", "Heredad creativa del diseño en Travesías"),
        ("INV-06", "Experimentación arquitectónica en Ciudad Abierta"),
    ]),
    ("SUB-04", [
        ("INV-03", "Borde costero y ciudad-puerto"),
    ]),
    ("SUB-05", []),
    ("SUB-06", [
        ("INV-10", "Arquitectura como medio didáctico"),
        ("INV-10", "Capacidad didáctica de la obra"),
    ]),
    ("SUB-07", [
        ("INV-12", "Comunicación aumentativa alternativa"),
    ]),
    ("SUB-08", [
        ("INV-01", "Diseño arquitectónico"),
        ("INV-17", "Diseño arquitectónico"),
    ]),
    ("SUB-09", [
        ("INV-01", "Diseño de espacios educativos"),
        ("INV-01", "Espacios educativos"),
        ("INV-01", "Arquitectura escolar"),
    ]),
    ("SUB-10", [
        ("INV-02", "Diseño de interacción y experiencia de usuarios"),
        ("INV-12", "Diseño de interacción y servicios"),
        ("INV-21", "Interacción como material de diseño"),
        ("INV-21", "Interacción física en espacios expositivos"),
    ]),
    ("SUB-11", [
        ("INV-12", "Diseño y vida independiente"),
    ]),
    ("SUB-12", [
        ("INV-15", "Mediación territorial y hospitalidad en lo habitable"),
    ]),
    ("SUB-13", []),
    ("SUB-14", [
        ("INV-10", "Espacios de aprendizaje en contextos socialmente vulnerables"),
    ]),
    ("SUB-15", [
        ("INV-10", "Estimulación temprana"),
    ]),
    ("SUB-16", [
        ("INV-10", "Espacios de aprendizaje en contextos socialmente vulnerables"),
    ]),
    ("SUB-17", []),
    ("SUB-18", [
        ("INV-19", "Innovación social y tecnológica"),
    ]),
    ("SUB-19", [
        ("INV-16", "Pedagogía del Diseño"),
        ("INV-16", "Diseño de información visual"),
    ]),
    ("SUB-20", [
        ("INV-02", "Metodologías y métodos de diseño"),
        ("INV-07", "Metodologías de investigación cualitativa"),
    ]),
    ("SUB-21", [
        ("INV-08", "Decolonialidad y urbanización"),
        ("INV-11", "Diseño decolonial"),
    ]),
    ("SUB-22", [
        ("INV-05", "Creación colectiva y experimentación arquitectónica"),
        ("INV-22", "Cooperativismo"),
        ("INV-22", "Movimientos autogestionados"),
    ]),
    ("SUB-23", []),
    ("SUB-24", []),
    ("SUB-25", [
        ("INV-11", "Diseño y artesanía"),
        ("INV-11", "Interacción entre oficio y cultura"),
    ]),
    ("SUB-26", [
        ("INV-01", "Teoría de la arquitectura"),
        ("INV-17", "Teoría de la arquitectura"),
        ("INV-18", "Historia y crítica de la arquitectura moderna"),
        ("INV-18", "Arquitectura moderna latinoamericana"),
    ]),
    ("SUB-27", [
        ("INV-15", "Transferencia tecnológica distribuida"),
        ("INV-19", "Transferencias disciplinares y formativas"),
    ]),
    ("SUB-28", [
        ("INV-22", "Caminar como práctica estética y artística"),
        ("INV-22", "Deriva urbana"),
    ]),
    ("SUB-29", [
        ("INV-08", "Urbanización capitalista y extractiva"),
        ("INV-08", "Ecología política y diseño de la transición"),
        ("INV-07", "Teoría urbana"),
    ]),
    ("SUB-30", [
        ("INV-03", "Borde costero y ciudad-puerto"),
    ]),
    # ----- Mapeos para sublíneas previamente huérfanas -----
    ("SUB-05", [
        # Hipótesis razonable: tema afín de Mercado (resiliencias)
        ("INV-08", "Resiliencias humanas y más-que-humanas"),
    ]),
    ("SUB-13", [
        ("INV-18", "Estudio del concepto de espacio"),
    ]),
    ("SUB-17", [
        ("INV-19", "Transferencias disciplinares y formativas"),
        ("INV-15", "Aprendizaje desde el hacer-crear"),
        ("INV-15", "Investigación proyectual"),
    ]),
    ("SUB-23", [
        ("INV-01", "Arquitectura escolar"),
        ("INV-01", "Espacios educativos"),
    ]),
    ("SUB-24", [
        ("INV-01", "Arquitectura escolar"),
        ("INV-10", "Capacidad didáctica de la obra"),
    ]),
    # ----- Mapeos para las 22 sublíneas nuevas -----
    ("SUB-31", [
        ("INV-07", "Vivienda en arriendo"),
        ("INV-07", "Financiarización de la vivienda"),
    ]),
    ("SUB-32", [
        ("INV-08", "Comunes y comunalidad"),
        ("INV-08", "Resiliencias humanas y más-que-humanas"),
        ("INV-08", "Transformaciones socioecológicas"),
        ("INV-08", "Frontera, relacionalidad y diferencia"),
        ("INV-21", "Multiterritorialidad y frontera"),
    ]),
    ("SUB-33", [
        ("INV-06", "Habitabilidad en infraestructura urbana"),
        ("INV-06", "Equipamiento e infraestructura urbana"),
        ("INV-06", "Movilidad sustentable"),
        ("INV-06", "Habitabilidad en infraestructura urbana técnica y natural"),
        ("INV-06", "Temas urbanos"),
        ("INV-06", "Ciudad y territorio"),
        ("INV-05", "Equipamientos de interpretación, educación y servicios"),
        ("INV-05", "Equipamiento cultural"),
        ("INV-03", "Movilidad y orden urbano en Valparaíso"),
        ("INV-03", "Muros de contención como estructura urbana"),
    ]),
    ("SUB-34", [
        ("INV-04", "Ciudad-teatro"),
        ("INV-04", "Cultura y ciudad"),
        ("INV-04", "Fenómenos sociales informales"),
        ("INV-04", "Apropiación del espacio público como acto escénico"),
        ("INV-04", "Espacio y forma arquitectónica"),
    ]),
    ("SUB-35", [
        ("INV-05", "Áreas naturales protegidas"),
        ("INV-05", "Relación artificio y naturaleza"),
        ("INV-05", "Patrimonio natural y cultural"),
        ("INV-05", "Sustentabilidad del patrimonio"),
        ("INV-05", "Arquitectura como interfaz entre habitante y naturaleza"),
        ("INV-05", "Áreas urbanas"),
    ]),
    ("SUB-36", [
        ("INV-03", "Patrimonio arquitectónico y territorial"),
        ("INV-03", "Rehabilitación patrimonial"),
        ("INV-03", "Frentes marítimos y ribereños"),
        ("INV-03", "Territorio marítimo y portuario"),
        ("INV-21", "Patrimonio urbano y memoria material"),
    ]),
    ("SUB-37", [
        ("INV-22", "Deriva urbana"),
        ("INV-22", "Urbanismo afectivo"),
        ("INV-22", "Reutilización del patrimonio abandonado"),
        ("INV-22", "Reuso del patrimonio público en abandono"),
    ]),
    ("SUB-38", [
        ("INV-22", "Investigación-acción"),
        ("INV-22", "Artes urbanas"),
    ]),
    ("SUB-39", [
        ("INV-14", "Inversión pública y evaluación social de proyectos"),
        ("INV-14", "Desarrollo de metodologías para la evaluación social de proyectos"),
        ("INV-14", "Valoración de beneficios sociales generados por proyectos de inversión"),
    ]),
    ("SUB-40", [
        ("INV-13", "Mobiliario y lugar"),
        ("INV-13", "Actos de la vida urbana y cotidiana"),
        ("INV-13", "Materialidad y obra"),
        ("INV-13", "Sistemas de proyecto desde origen a implementación"),
        ("INV-13", "Habitabilidad"),
    ]),
    ("SUB-41", [
        # Jeldes — núcleo: fabricación digital y paramétrico
        ("INV-15", "Fabricación digital"),
        ("INV-15", "Fablabs"),
        ("INV-15", "Modelado paramétrico"),
        ("INV-15", "Modelado paramétrico interpretativo"),
        ("INV-15", "Diseño tecnológico aplicado"),
    ]),
    ("SUB-42", [
        ("INV-12", "Diseño para la democracia"),
        ("INV-12", "Plataformas de comunicación ciudadana"),
    ]),
    ("SUB-43", [
        ("INV-12", "Comunicación visual"),
        ("INV-16", "Diseño de información visual"),
        ("INV-20", "Diseño gráfico editorial"),
        ("INV-20", "Montajes expositivos"),
        ("INV-20", "Contenidos patrimoniales y su exposición material"),
        ("INV-21", "Interacción física en espacios expositivos"),
        ("INV-21", "Visualización territorial"),
    ]),
    ("SUB-44", [
        ("INV-11", "Diseño social y territorial"),
        ("INV-11", "Objetos, contexto social y cultural"),
        ("INV-11", "Patrimonio cultural material e inmaterial"),
    ]),
    ("SUB-45", [
        ("INV-02", "Confort térmico y bienestar"),
    ]),
    ("SUB-46", [
        ("INV-19", "Máquinas expresivas"),
        ("INV-19", "Algoritmos cinéticos"),
        ("INV-21", "Estructuras cinéticas y mecánico-expresivas"),
        ("INV-21", "Formas resistentes del diseño"),
        ("INV-21", "Taller topológico multiescalar"),
    ]),
    ("SUB-47", [
        ("INV-15", "Tecnología y sociedad"),
        ("INV-19", "Techné y mundo"),
        ("INV-19", "Epistemología del diseño"),
        ("INV-19", "Observación como transmisión de ethos"),
        ("INV-19", "Habitabilidad y biodiversidad"),
        ("INV-05", "Proyecto y obra como generación de conocimiento"),
    ]),
    ("SUB-48", [
        ("INV-17", "Arte y arquitectura"),
    ]),
    ("SUB-49", [
        ("INV-18", "Historia y crítica de la arquitectura moderna"),
        ("INV-18", "Arquitectura moderna latinoamericana"),
        ("INV-18", "Habitar moderno"),
        ("INV-18", "Arquitectura del Cono Sur"),
    ]),
    ("SUB-50", [
        ("INV-09", "Relación poesía-oficio"),
        ("INV-09", "Seminario de América"),
        ("INV-20", "Poética del oficio"),
        ("INV-20", "Actos poéticos"),
    ]),
    ("SUB-51", [
        ("INV-20", "Fundamentos poéticos y artísticos del diseño"),
        ("INV-20", "Heredad creativa del diseño en Travesías"),
    ]),
    ("SUB-52", [
        ("INV-01", "Reforma escolar de 1965 en Chile"),
        ("INV-01", "Relación entre arquitectura y aprendizaje"),
        ("INV-17", "Educación en arquitectura"),
        ("INV-09", "Investigación en posgrado dentro de Educación, Espacio y Aprendizaje"),
    ]),
    ("SUB-53", [
        # Spencer — IA y sistemas inteligentes orientados a inclusión y democracia
        ("INV-12", "Diseño de sistemas inteligentes"),
        ("INV-12", "IA como material de diseño"),
    ]),
]

SUB_TEMA = []
for subid, refs in SUB_TEMA_MAPPINGS:
    for inv_id, tema_text in refs:
        tid = find_tema(tema_text, inv_id)
        if tid:
            SUB_TEMA.append((subid, tid))

# ---------------------------------------------------------------------------
# Sello formativo del Doctorado (hoja 17_Sello)
# Variante 1 = ELEGIDA. Resto se conservan para trazabilidad.
# Regla de estilo del programa: variante elegida sin negaciones ni em-dashes.
# ---------------------------------------------------------------------------
SELLO_VARIANTES = [
    # (variante_id, foco, texto, elegido)
    (
        1,
        "La obra como argumento (ELEGIDO)",
        "El doctorado forma investigadores para quienes la obra es origen y prueba de la tesis. "
        "Esa obra, sea edificada, fabricada, escrita o dibujada, vale como argumento por sí misma, "
        "y el discurso doctoral se construye para hacerla legible y discutible. "
        "La pregunta que esa obra está llamada a argumentar es la que el programa hace suya: "
        "cómo reinventar el habitar humano. "
        "Las cuatro líneas del doctorado aportan los principios que esa pregunta convoca: "
        "la inclusión y la autonomía como condiciones de quien habita, "
        "las ecologías del territorio como condiciones de dónde, "
        "la herencia disciplinar de la e[ad] y la teoría del proyecto como condiciones desde dónde se piensa, "
        "y el oficio y los medios técnicos como condiciones con qué se hace. "
        "Doctorarse aquí es haber producido una obra capaz de sostener esos principios al ponerse a prueba "
        "y, en su mejor versión, capaz de reformularlos.",
        True,
    ),
    (
        2,
        "Palabra y obra inseparables",
        "Lo que aquí se forma es la imposibilidad de separar lo que se dice de lo que se hace: "
        "cada una porta el argumento de la otra. El doctorando se gradúa cuando ese vínculo deja "
        "de ser declarativo y aparece, demostrable, en su trabajo.",
        False,
    ),
    (
        3,
        "Tres disciplinas encadenadas",
        "La formación doctoral está organizada en torno a tres disciplinas que se exigen entre sí: "
        "una observación sostenida en el tiempo, una palabra precisa para nombrar lo observado, "
        "y una obra que las pone a prueba. La tesis doctoral es haber recorrido ese encadenamiento "
        "al menos una vez con resultado defendible.",
        False,
    ),
    (
        4,
        "Pensar con el material",
        "Se forma una capacidad poco frecuente: pensar con el material y no sobre el material. "
        "Plano, maqueta, terreno, prototipo y página son los lugares donde el pensamiento ocurre, "
        "no representaciones de un pensamiento que ocurrió en otra parte.",
        False,
    ),
    (
        5,
        "Responsabilidad por la obra propia",
        "El doctorando responde por una obra que hace, no por una literatura que comenta. "
        "El programa no forma críticos de arquitectura y diseño: forma investigadores que producen "
        "obra y la sostienen en el plano académico.",
        False,
    ),
    (
        6,
        "Travesía como exigencia",
        "La formación incorpora la travesía como condición: el continente se piensa cruzándolo, "
        "observándolo y dejando obra en él. Quien se doctora aquí no escribe sobre territorios que "
        "no ha recorrido ni sobre comunidades con las que no ha trabajado.",
        False,
    ),
]

# ---------------------------------------------------------------------------
# Proximidad temática (hoja 18_Proximidad_Tematica)
# Pares simétricos a < b. Afinidad 0..1. Inferido inicialmente; afinar a mano.
# Cada par implica afinidad inversa equivalente (b↔a).
# ---------------------------------------------------------------------------
PROXIMIDAD_PARES = [
    # (sublinea_a, sublinea_b, afinidad, razonamiento)
    # Cluster Accesibilidad / Inclusión / Vida independiente / Interacción
    ("SUB-01", "SUB-02", 0.85, "La accesibilidad cognitiva es un caso específico del marco de accesibilidad e inclusión."),
    ("SUB-01", "SUB-07", 0.80, "CAA es la aplicación operativa más directa de la accesibilidad cognitiva."),
    ("SUB-01", "SUB-09", 0.40, "Bridge: accesibilidad cognitiva tiene aplicación clara en espacios educativos."),
    ("SUB-01", "SUB-10", 0.65, "Codiseño es práctica de diseño de interacción centrado en personas."),
    ("SUB-01", "SUB-11", 0.70, "Codiseño orientado a sostener la autonomía y vida independiente."),
    ("SUB-01", "SUB-15", 0.55, "Estimulación temprana opera con principios de accesibilidad cognitiva infantil."),
    ("SUB-01", "SUB-20", 0.55, "El codiseño es un método de diseño consolidado."),
    ("SUB-02", "SUB-07", 0.65, "CAA es la dimensión comunicacional de la inclusión."),
    ("SUB-02", "SUB-10", 0.50, "Inclusión cruza la práctica de servicios y UX."),
    ("SUB-02", "SUB-11", 0.75, "Vida independiente es el horizonte de la inclusión."),
    ("SUB-02", "SUB-12", 0.55, "La hospitalidad como categoría de inclusión radical."),
    ("SUB-02", "SUB-14", 0.50, "Aprendizaje en contextos vulnerables es inclusión educativa."),
    ("SUB-02", "SUB-16", 0.50, "Espacios educativos vulnerables como inclusión espacial."),
    ("SUB-02", "SUB-21", 0.45, "Decolonialidad amplía la inclusión a la diferencia cultural."),
    ("SUB-07", "SUB-10", 0.55, "CAA modela formas específicas de interacción."),
    ("SUB-07", "SUB-11", 0.65, "CAA es condición de comunicación para la vida independiente."),
    ("SUB-10", "SUB-11", 0.55, "Servicios y UX deben sostener la autonomía."),
    ("SUB-10", "SUB-20", 0.50, "Métodos de diseño se aplican intensivamente al diseño de interacción."),
    ("SUB-11", "SUB-12", 0.45, "Hospitalidad como condición de la vida con otros."),
    ("SUB-11", "SUB-18", 0.40, "Innovación social puede orientarse a sistemas para la vida independiente."),

    # Cluster Riesgo costero / Desastres / Vulnerabilidad
    ("SUB-04", "SUB-05", 0.85, "Riesgos costeros y desastres comparten marco analítico y operativo."),
    ("SUB-04", "SUB-29", 0.55, "Adaptabilidad urbana costera entra en diálogo con ecología política urbana."),
    ("SUB-04", "SUB-30", 0.90, "Adaptabilidad y vulnerabilidad costera son caras complementarias del mismo problema."),
    ("SUB-05", "SUB-08", 0.50, "Adaptación formal ante desastres es una respuesta del diseño arquitectónico."),
    ("SUB-05", "SUB-29", 0.50, "Desastres están entrelazados con condiciones políticas y ecológicas."),
    ("SUB-05", "SUB-30", 0.80, "Vulnerabilidad costera es la causa que la adaptación formal aborda."),
    ("SUB-08", "SUB-30", 0.40, "Diseño arquitectónico se confronta con vulnerabilidad costera en proyectos."),
    ("SUB-29", "SUB-30", 0.55, "Vulnerabilidad costera es leída desde la urbanización política."),

    # Cluster Educación / Espacios escolares / Enseñanza-aprendizaje
    ("SUB-06", "SUB-09", 0.70, "Arquitectura como medio didáctico se realiza en el diseño de espacios educativos."),
    ("SUB-06", "SUB-14", 0.50, "El medio didáctico tiene aplicación en aprendizajes vulnerables."),
    ("SUB-06", "SUB-17", 0.45, "Pensamiento creativo se forma con la obra como medio didáctico."),
    ("SUB-06", "SUB-19", 0.75, "Medios de aprendizaje son el correlato instrumental del medio didáctico."),
    ("SUB-06", "SUB-23", 0.55, "Programas educativos se articulan con el medio didáctico arquitectónico."),
    ("SUB-06", "SUB-24", 0.75, "El proyecto escolar es la forma máxima del medio didáctico."),
    ("SUB-06", "SUB-26", 0.45, "Teoría arquitectónica fundamenta la lectura didáctica de la obra."),
    ("SUB-09", "SUB-14", 0.75, "Aprendizaje en contextos vulnerables exige espacios educativos pensados."),
    ("SUB-09", "SUB-15", 0.65, "Estimulación temprana requiere diseño específico de espacio."),
    ("SUB-09", "SUB-16", 0.85, "Espacios educativos vulnerables es subdominio directo."),
    ("SUB-09", "SUB-19", 0.55, "Medios de aprendizaje habitan espacios educativos."),
    ("SUB-09", "SUB-23", 0.85, "Programas educativos definen los requisitos del espacio."),
    ("SUB-09", "SUB-24", 0.80, "Proyecto escolar es el caso paradigmático del diseño educativo."),
    ("SUB-14", "SUB-15", 0.65, "Estimulación temprana ocurre frecuentemente en contextos vulnerables."),
    ("SUB-14", "SUB-16", 0.90, "Aprendizaje y espacios educativos en contextos vulnerables son lecturas casi solapadas."),
    ("SUB-14", "SUB-19", 0.55, "Medios de aprendizaje se ajustan a contextos vulnerables."),
    ("SUB-15", "SUB-16", 0.65, "Estimulación temprana suele desarrollarse en contextos vulnerables."),
    ("SUB-16", "SUB-23", 0.65, "Programas educativos definen los espacios para contextos vulnerables."),
    ("SUB-19", "SUB-23", 0.55, "Medios de aprendizaje son herramientas operativas de los programas educativos."),
    ("SUB-23", "SUB-24", 0.80, "Programas educativos sustentan la pregunta del proyecto de arquitectura escolar."),

    # Cluster Acervo / Teoría / Heritage e[ad]
    ("SUB-03", "SUB-08", 0.45, "Acervo histórico ilumina la comprensión del diseño arquitectónico."),
    ("SUB-03", "SUB-12", 0.65, "La hospitalidad es categoría central del acervo de la EAD."),
    ("SUB-03", "SUB-13", 0.65, "El vacío arquitectónico es concepto formado en el acervo de la Escuela."),
    ("SUB-03", "SUB-19", 0.40, "Medios de aprendizaje pueden anclarse en el acervo de la Escuela."),
    ("SUB-03", "SUB-21", 0.40, "Decolonialidad entra en diálogo con la lectura americana del acervo."),
    ("SUB-03", "SUB-26", 0.75, "Teoría e historia se nutren del acervo institucional."),
    ("SUB-03", "SUB-28", 0.70, "Travesías son parte central del acervo histórico de la Escuela."),
    ("SUB-08", "SUB-13", 0.70, "El vacío arquitectónico es categoría central del diseño arquitectónico."),
    ("SUB-08", "SUB-20", 0.55, "Métodos de diseño se aplican al diseño arquitectónico."),
    ("SUB-08", "SUB-24", 0.65, "Proyecto escolar es un caso de diseño arquitectónico."),
    ("SUB-08", "SUB-26", 0.65, "Teoría e historia fundamentan el diseño arquitectónico."),
    ("SUB-08", "SUB-28", 0.50, "Travesías han sido escenario de proyectos arquitectónicos."),
    ("SUB-12", "SUB-13", 0.50, "Hospitalidad y vacío comparten registro fenomenológico de la EAD."),
    ("SUB-12", "SUB-22", 0.45, "Hospitalidad es condición de las prácticas colectivas."),
    ("SUB-12", "SUB-28", 0.60, "Hospitalidad y travesías están enlazadas en la tradición de Amereida."),
    ("SUB-13", "SUB-26", 0.70, "El vacío es categoría teórico-arquitectónica de raíz e[ad]."),
    ("SUB-13", "SUB-28", 0.55, "Travesías son experiencia de habitar el vacío continental."),
    ("SUB-24", "SUB-26", 0.40, "El proyecto escolar tiene una historia y teoría disciplinar."),
    ("SUB-26", "SUB-28", 0.50, "Travesías como objeto de teoría e historia de la arquitectura."),

    # Cluster Métodos / Técnicas / Oficio / Industrias creativas
    ("SUB-17", "SUB-18", 0.40, "Pensamiento y acción creativa puede orientar la innovación social."),
    ("SUB-17", "SUB-19", 0.50, "Pensamiento creativo se transfiere por medios de aprendizaje específicos."),
    ("SUB-17", "SUB-20", 0.65, "Métodos de diseño formalizan el pensamiento y acción creativa."),
    ("SUB-17", "SUB-25", 0.50, "Saberes técnicos análogos sostienen la formación creativa."),
    ("SUB-17", "SUB-27", 0.45, "Transferencia tecnológica encuentra cauce en formación creativa."),
    ("SUB-18", "SUB-20", 0.45, "Métodos de diseño habilitan la innovación social."),
    ("SUB-18", "SUB-22", 0.50, "Innovación social se entrelaza con prácticas colectivas."),
    ("SUB-18", "SUB-25", 0.55, "Saberes técnicos análogos sostienen industrias creativas locales."),
    ("SUB-18", "SUB-27", 0.75, "Transferencia tecnológica es vector clave del emprendimiento local."),
    ("SUB-18", "SUB-29", 0.40, "Innovación local entra en diálogo con la urbanización política."),
    ("SUB-20", "SUB-25", 0.55, "Saberes técnicos análogos integran el repertorio de los métodos."),
    ("SUB-20", "SUB-27", 0.40, "Transferencia tecnológica usa métodos de diseño formalizados."),
    ("SUB-25", "SUB-27", 0.80, "Saberes técnicos análogos y transferencia tecnológica son caras del mismo programa."),

    # Cluster Urbanismo / Decolonial / Colectivo
    ("SUB-21", "SUB-22", 0.70, "Prácticas colectivas son una vía decolonial de producción del habitar."),
    ("SUB-21", "SUB-28", 0.65, "Travesías y Amereida son matriz histórica del giro decolonial e[ad]."),
    ("SUB-21", "SUB-29", 0.75, "Ecología política urbana lee la urbanización en clave decolonial."),
    ("SUB-22", "SUB-28", 0.55, "Travesías son una práctica colectiva paradigmática."),
    ("SUB-22", "SUB-29", 0.65, "Prácticas colectivas ocurren en y sobre la urbanización."),

    # Cross-cluster Acervo ↔ Oficio
    ("SUB-03", "SUB-25", 0.50, "Saberes técnicos análogos están heredados desde el acervo institucional."),
]

# Validación de invariantes a < b y unicidad
_seen_pares = set()
for a, b, *_ in PROXIMIDAD_PARES:
    assert a < b, f"Par no ordenado: {a} >= {b} (debe ser a < b)"
    key = (a, b)
    assert key not in _seen_pares, f"Par duplicado: {key}"
    _seen_pares.add(key)

# Mapa SUB-ID -> nombre para enriquecer la hoja Sublinea_Tema
SUB_NOMBRE = {sid: nombre for sid, nombre, *_ in SUBLINEAS}

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="111827")
NOTE_FONT = Font(name="Arial", italic=True, color="6B7280", size=10)
DRAFT_FILL = PatternFill("solid", start_color="FEF3C7")  # amarillo suave para borrador
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_table(ws, headers, rows, widths, title=None, note=None, highlight_col=None, highlight_value=None):
    r = 1
    if title:
        ws.cell(row=r, column=1, value=title).font = TITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        r += 1
    if note:
        ws.cell(row=r, column=1, value=note).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        r += 1
        r += 1  # blank row
    header_row = r
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers))
    for row in rows:
        r += 1
        for i, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Arial", size=11)
            cell.border = BORDER
            if highlight_col is not None and highlight_value is not None:
                if isinstance(row[highlight_col], str) and highlight_value in row[highlight_col]:
                    cell.fill = DRAFT_FILL
    autosize(ws, widths)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return header_row


# ---------------------------------------------------------------------------
# Construir el workbook
# ---------------------------------------------------------------------------
wb = Workbook()

# 00_Lectura
ws = wb.active
ws.title = "00_Lectura"
ws["A1"] = "MAD-map · Estructura de datos v2"
ws["A1"].font = Font(name="Arial", bold=True, size=18, color="111827")
ws["A2"] = "Régimen sin áreas-MAD: Líneas troncales sostenidas por Laboratorios + sublíneas + temas."
ws["A2"].font = NOTE_FONT
ws["A4"] = "Modelo relacional"
ws["A4"].font = Font(name="Arial", bold=True, size=12)
hierarchy_rows = [
    ("Línea (4 troncales)", "Estructura superior. Lo que la Escuela declara y sostiene. BORRADOR."),
    ("Sublínea (~30)", "Pertenece a UNA sola Línea. Granularidad operativa. Banco depurado."),
    ("Tema declarado (~150)", "Tema textual del perfil de un investigador. Puede colgar de varias Sublíneas (m:n)."),
    ("Investigador", "Profesor de la EAD-PUCV. Ancla los Temas. Afilia a Áreas, Labs y Modos."),
    ("Área del Postgrado", "ECH / EAA / FCT. Eje ortogonal — marco amplio, no contenedor jerárquico."),
    ("Modo de investigar", "Historiografía / Teoría crítica / Investigación proyectual. Atributo del Investigador y de la Línea."),
    ("Laboratorio (5)", "Grupo asociativo. Sostiene 1+ Líneas (m:n). Operacionaliza Salidas."),
    ("Salida", "Industria / Academia / Estado. Producida por Laboratorios."),
]
for i, (label, desc) in enumerate(hierarchy_rows, start=5):
    ws.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=i, column=2, value=desc).font = Font(name="Arial", size=11)
    ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 110

ws["A14"] = "Convenciones"
ws["A14"].font = Font(name="Arial", bold=True, size=12)
notes = [
    ("BORRADOR (fondo amarillo)", "Filas/celdas con valores propuestos por revisar. La Línea-04 y los mapeos m:n están sujetos a curaduría."),
    ("IDs", "LIN-NN, SUB-NN, TEM-NNN, INV-NN, LAB-XXX, MOD-NN, SAL-XXX. Estables — no renombrar tras enlazar."),
    ("Cardinalidad Sublínea→Línea", "1:N estricta (una Sublínea pertenece a UNA Línea). Cambia el linea_id si el ajuste lo amerita."),
    ("Cardinalidad Tema→Sublínea", "M:N (un Tema puede colgar de varias Sublíneas). Editar en la hoja 09_Sublinea_Tema."),
    ("Cardinalidad Lab→Línea", "M:N. Un Lab puede sostener varias Líneas. Editar en 10_Lab_Linea."),
    ("Decisiones pendientes", "Hoja 15_Decisiones — preguntas abiertas que afectan el modelo."),
]
for i, (label, desc) in enumerate(notes, start=15):
    ws.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=i, column=2, value=desc).font = Font(name="Arial", size=11)
    ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")

# 01_Líneas
ws = wb.create_sheet("01_Lineas")
write_table(
    ws,
    headers=["id", "nombre", "descripción", "estado"],
    rows=[(l[0], l[1], l[2], l[3]) for l in LINEAS],
    widths=[12, 55, 95, 26],
    title="01 · Líneas troncales (4 a definir — BORRADOR)",
    note="Cada Sublínea pertenece a una Línea. Los Laboratorios pueden sostener una o varias Líneas. Borrador derivado de la hipótesis de compactación.",
    highlight_col=3,
    highlight_value="BORRADOR",
)

# 02_Sublíneas
ws = wb.create_sheet("02_Sublineas")
write_table(
    ws,
    headers=["id", "nombre", "linea_id", "area_id", "notas"],
    rows=SUBLINEAS,
    widths=[10, 55, 12, 10, 50],
    title="02 · Sublíneas (banco depurado, ~30)",
    note="Cada sublínea pertenece a UNA línea troncal. La asignación linea_id es BORRADOR — revisar caso por caso.",
)

# 03_Áreas
ws = wb.create_sheet("03_Areas")
write_table(
    ws,
    headers=["código", "nombre", "descripción"],
    rows=AREAS,
    widths=[10, 50, 110],
    title="03 · Áreas del Postgrado",
    note="Eje ortogonal a las Líneas. Marcos amplios, no compartimentos.",
)

# 04_Modos
ws = wb.create_sheet("04_Modos")
write_table(
    ws,
    headers=["id", "nombre", "descripción"],
    rows=MODOS,
    widths=[10, 30, 100],
    title="04 · Modos de investigar",
)

# 05_Salidas
ws = wb.create_sheet("05_Salidas")
write_table(
    ws,
    headers=["id", "nombre", "descripción"],
    rows=SALIDAS,
    widths=[10, 20, 100],
    title="05 · Salidas (canales de transferencia)",
    note="Los Laboratorios son la unidad operativa que activa las Salidas.",
)

# 06_Laboratorios
ws = wb.create_sheet("06_Laboratorios")
write_table(
    ws,
    headers=["id", "nombre", "descripción"],
    rows=LABORATORIOS,
    widths=[12, 38, 100],
    title="06 · Laboratorios (grupos asociativos, nombres tentativos)",
)

# 07_Investigadores
ws = wb.create_sheet("07_Investigadores")
write_table(
    ws,
    headers=["id", "nombre", "área_principal", "perfil_casiopea", "estado_perfil"],
    rows=INVESTIGADORES,
    widths=[10, 28, 14, 60, 38],
    title="07 · Investigadores (cuerpo académico EAD-PUCV)",
    note="Área principal es lectura — un investigador puede operar en varias. Profundización en 12_Investigador_Lab y 13_Investigador_Modo.",
)

# 08_Temas
ws = wb.create_sheet("08_Temas")
write_table(
    ws,
    headers=["id", "investigador_id", "tema"],
    rows=TEMAS,
    widths=[10, 14, 100],
    title="08 · Temas declarados (Casiopea + ANID)",
    note="Cada Tema pertenece a UN Investigador. Las relaciones a Sublíneas viven en 09_Sublinea_Tema (m:n).",
)

# 09_Sublinea_Tema (m:n)
ws = wb.create_sheet("09_Sublinea_Tema")
write_table(
    ws,
    headers=["sublinea_id", "tema_id"],
    rows=SUB_TEMA,
    widths=[15, 15],
    title="09 · Sublínea ↔ Tema (m:n)",
    note="Mapeo de alta confianza. Sublíneas sin filas aquí están pendientes de poblar con temas declarados.",
)

# 10_Lab_Linea (m:n)
ws = wb.create_sheet("10_Lab_Linea")
write_table(
    ws,
    headers=["lab_id", "linea_id"],
    rows=LAB_LINEA,
    widths=[14, 14],
    title="10 · Laboratorio ↔ Línea (m:n)",
    note="Un Lab sostiene una o varias Líneas. BORRADOR.",
)

# 11_Lab_Salida (m:n)
ws = wb.create_sheet("11_Lab_Salida")
write_table(
    ws,
    headers=["lab_id", "salida_id"],
    rows=LAB_SALIDA,
    widths=[14, 14],
    title="11 · Laboratorio ↔ Salida (m:n)",
    note="Los Labs operacionalizan las Salidas. BORRADOR.",
)

# 12_Investigador_Lab (m:n)
ws = wb.create_sheet("12_Investigador_Lab")
write_table(
    ws,
    headers=["investigador_id", "lab_id"],
    rows=INV_LAB,
    widths=[18, 14],
    title="12 · Investigador ↔ Laboratorio (m:n)",
    note="Lectura desde el material. BORRADOR — algunos investigadores podrían estar en más laboratorios de los registrados aquí.",
)

# 13_Investigador_Modo (m:n)
ws = wb.create_sheet("13_Investigador_Modo")
write_table(
    ws,
    headers=["investigador_id", "modo_id"],
    rows=INV_MODO,
    widths=[18, 12],
    title="13 · Investigador ↔ Modo (m:n)",
    note="Lectura desde el material — modos predominantes declarados o inferidos.",
)

# 14_Linea_Modo (m:n) — modos predominantes por línea (vacío para curaduría)
ws = wb.create_sheet("14_Linea_Modo")
write_table(
    ws,
    headers=["linea_id", "modo_id", "nivel"],
    rows=[
        ("LIN-01", "MOD-03", "predominante"),
        ("LIN-01", "MOD-02", "presente"),
        ("LIN-02", "MOD-02", "predominante"),
        ("LIN-02", "MOD-03", "presente"),
        ("LIN-03", "MOD-01", "predominante"),
        ("LIN-03", "MOD-02", "predominante"),
        ("LIN-04", "MOD-03", "predominante"),
    ],
    widths=[12, 12, 18],
    title="14 · Línea ↔ Modo (m:n)",
    note="Modos predominantes por Línea — útil para coloreo/filtros del grafo. BORRADOR.",
)

# 15_Decisiones_pendientes
ws = wb.create_sheet("15_Decisiones")
decisiones = [
    ("D-01", "¿3 o 4 Líneas troncales?",
     "El brief inicial dice 4. En la elicitación apareció '3 líneas x definir'. Se trabaja con 4 BORRADOR; la 4ª (Métodos, técnicas y medios) es candidata a fusionar con LIN-01 o LIN-03 si se decide compactar a 3.",
     "Abierta"),
    ("D-02", "Nombres definitivos de las 4 Líneas",
     "Los nombres actuales son derivados de la hipótesis de compactación. Confirmar redacción institucional (la que iría a CNA).",
     "Abierta"),
    ("D-03", "Áreas del Postgrado: ¿quedan? ¿se renombran?",
     "Material plantea variantes (ej. 'Extensión, Territorio y Ecologías del Habitar'). Si se mantienen como eje ortogonal, decidir nombre canónico.",
     "Abierta"),
    ("D-04", "Sublínea SUB-22 'Prácticas colectivas'",
     "Asignada provisoriamente a LIN-02. Tiene respaldo también en LIN-03. Decidir hogar.",
     "Abierta"),
    ("D-05", "Sublínea SUB-21 'Perspectivas decoloniales'",
     "Asignada provisoriamente a LIN-02 (ecologías). Podría leerse en LIN-01 (inclusión) o quedar transversal.",
     "Abierta"),
    ("D-06", "Cobertura de mapeo Sublínea↔Tema",
     "Mapeo solo de alta confianza. Sublíneas SUB-05, SUB-13, SUB-17, SUB-23, SUB-24 sin temas asociados — completar o decidir si retiran del banco.",
     "Abierta"),
    ("D-07", "Proximidad semántica latente",
     "Pendiente definir: ¿es declarada (curaduría humana, valor de afinidad) o computada (embeddings sobre nombres+descripciones)? Esto cambia los datos que la planilla debe contener (vector_id, score, etc.).",
     "Abierta"),
    ("D-08", "¿Las Salidas se asocian solo a Labs, o también a Líneas/Sublíneas?",
     "Hoy el modelo asocia Salida↔Lab. Confirmar si una Línea declara una Salida directamente, o si siempre pasa por su(s) Lab(s).",
     "Abierta"),
    ("D-09", "Investigadores afiliados a más de un Área",
     "Modelo actual usa 'área_principal' (1:1). Si hay investigadores con afiliación múltiple, conviene una hoja Investigador↔Área (m:n).",
     "Abierta"),
    ("D-10", "Fuente única de verdad",
     "Esta planilla es ahora la fuente. Decidir cómo se sincroniza con la planilla anterior (lineas-areas-mad.csv) y si el grafo del index.html consume desde aquí.",
     "Abierta"),
]
write_table(
    ws,
    headers=["id", "decisión", "contexto", "estado"],
    rows=decisiones,
    widths=[8, 60, 100, 14],
    title="15 · Decisiones pendientes",
    note="Cada decisión bloquea o refina el modelo. Marcar como 'Resuelta' con la respuesta.",
)
# Pintar todas las decisiones como BORRADOR
for r in range(5, 5 + len(decisiones)):
    for c in range(1, 5):
        ws.cell(row=r, column=c).fill = DRAFT_FILL

# 16_Resumen (vista de control)
ws = wb.create_sheet("16_Resumen")
ws["A1"] = "Resumen cuantitativo"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Conteos por entidad (auto-calculados)"
ws["A3"].font = Font(name="Arial", bold=True, size=12)
counts = [
    ("Líneas",          "=COUNTA('01_Lineas'!A:A)-1"),
    ("Sublíneas",       "=COUNTA('02_Sublineas'!A:A)-1"),
    ("Áreas",           "=COUNTA('03_Areas'!A:A)-1"),
    ("Modos",           "=COUNTA('04_Modos'!A:A)-1"),
    ("Salidas",         "=COUNTA('05_Salidas'!A:A)-1"),
    ("Laboratorios",    "=COUNTA('06_Laboratorios'!A:A)-1"),
    ("Investigadores",  "=COUNTA('07_Investigadores'!A:A)-1"),
    ("Temas",           "=COUNTA('08_Temas'!A:A)-1"),
    ("Sublínea↔Tema",   "=COUNTA('09_Sublinea_Tema'!A:A)-1"),
    ("Lab↔Línea",       "=COUNTA('10_Lab_Linea'!A:A)-1"),
    ("Lab↔Salida",      "=COUNTA('11_Lab_Salida'!A:A)-1"),
    ("Investigador↔Lab","=COUNTA('12_Investigador_Lab'!A:A)-1"),
    ("Investigador↔Modo","=COUNTA('13_Investigador_Modo'!A:A)-1"),
    ("Línea↔Modo",      "=COUNTA('14_Linea_Modo'!A:A)-1"),
]
ws.cell(row=4, column=1, value="entidad / relación").font = HEADER_FONT
ws.cell(row=4, column=1).fill = HEADER_FILL
ws.cell(row=4, column=2, value="conteo").font = HEADER_FONT
ws.cell(row=4, column=2).fill = HEADER_FILL
for i, (label, formula) in enumerate(counts, start=5):
    ws.cell(row=i, column=1, value=label).font = Font(name="Arial", size=11)
    ws.cell(row=i, column=2, value=formula).font = Font(name="Arial", size=11)
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=2).border = BORDER

# Sección: cobertura
start = 5 + len(counts) + 2
ws.cell(row=start, column=1, value="Cobertura de mapeos").font = Font(name="Arial", bold=True, size=12)
ws.cell(row=start + 1, column=1, value="Temas con al menos una Sublínea")
ws.cell(row=start + 1, column=2, value="=SUMPRODUCT(1/COUNTIF('09_Sublinea_Tema'!B2:B500,'09_Sublinea_Tema'!B2:B500))")
ws.cell(row=start + 2, column=1, value="Total temas")
ws.cell(row=start + 2, column=2, value="=COUNTA('08_Temas'!A:A)-1")
ws.cell(row=start + 3, column=1, value="% temas mapeados")
ws.cell(row=start + 3, column=2, value=f"=B{start+1}/B{start+2}")
ws.cell(row=start + 3, column=2).number_format = "0.0%"

ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 20

# 17_Sello — sello formativo del doctorado
ws = wb.create_sheet("17_Sello")
CHOSEN_FILL = PatternFill("solid", start_color="DCFCE7")  # verde suave para el elegido
ALT_FILL    = PatternFill("solid", start_color="F3F4F6")  # gris suave para alternativas

ws["A1"] = "17 · Sello formativo del doctorado"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:C1")
ws["A2"] = ("Afirmación identitaria sobre qué se forma. La fila marcada (ELEGIDO, verde) "
            "es la versión que se usa para copy institucional. Las otras se conservan como "
            "trazabilidad de las alternativas exploradas. Regla de estilo de la elegida: "
            "redacción en afirmativo, sin em-dashes.")
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A2:C2")
ws.row_dimensions[2].height = 45

for i, h in enumerate(["Variante", "Foco", "Texto"], start=1):
    c = ws.cell(row=4, column=i, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER

for idx, (variante, foco, texto, elegido) in enumerate(SELLO_VARIANTES):
    r = 5 + idx
    fill = CHOSEN_FILL if elegido else ALT_FILL
    weight = elegido
    cells = [
        ws.cell(row=r, column=1, value=variante),
        ws.cell(row=r, column=2, value=foco),
        ws.cell(row=r, column=3, value=texto),
    ]
    for c in cells:
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(name="Arial", size=11, bold=weight)
        c.border = BORDER
    ws.row_dimensions[r].height = 220 if elegido else 70

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 110
ws.freeze_panes = "A5"

# 18_Proximidad_Tematica — matriz de afinidad declarada Sublínea↔Sublínea
ws = wb.create_sheet("18_Proximidad_Tematica")
HIGH_FILL = PatternFill("solid", start_color="DCFCE7")  # verde para afinidad >= 0.75

ws["A1"] = "18 · Proximidad temática (Sublínea ↔ Sublínea)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")

ws["A2"] = ("Matriz de afinidad declarada como pares simétricos. Inferida desde nombres + descripciones; "
            "estado=INFERIDO. Afinar a mano: corregir afinidad, agregar/quitar pares, marcar estado=CONFIRMADO/AJUSTADO/DESCARTADO. "
            "Cada par implica afinidad inversa equivalente (b↔a). Si re-generas la planilla con build_xlsx.py, "
            "los cambios manuales se sobrescriben — reflejarlos en PROXIMIDAD_PARES del script para que persistan.")
ws["A2"].font = NOTE_FONT
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A2:G2")
ws.row_dimensions[2].height = 60

prox_headers = ["sublinea_a_id", "sublinea_a_nombre", "sublinea_b_id", "sublinea_b_nombre",
                "afinidad", "razonamiento", "estado"]
for i, h in enumerate(prox_headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER

# Ordenar por afinidad descendente (más fuertes arriba)
PROX_SORTED = sorted(PROXIMIDAD_PARES, key=lambda p: (-p[2], p[0], p[1]))
for idx, (a, b, af, raz) in enumerate(PROX_SORTED):
    r = 5 + idx
    cells = [
        ws.cell(row=r, column=1, value=a),
        ws.cell(row=r, column=2, value=SUB_NOMBRE.get(a, a)),
        ws.cell(row=r, column=3, value=b),
        ws.cell(row=r, column=4, value=SUB_NOMBRE.get(b, b)),
        ws.cell(row=r, column=5, value=af),
        ws.cell(row=r, column=6, value=raz),
        ws.cell(row=r, column=7, value="INFERIDO"),
    ]
    for c in cells:
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(name="Arial", size=10)
        c.border = BORDER
    fill = HIGH_FILL if af >= 0.75 else DRAFT_FILL
    for c in cells:
        c.fill = fill
    cells[4].number_format = "0.00"

for i, w in enumerate([14, 38, 14, 38, 11, 70, 14], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:G{4 + len(PROX_SORTED)}"

dv = DataValidation(
    type="list",
    formula1='"INFERIDO,CONFIRMADO,AJUSTADO,DESCARTADO"',
    allow_blank=False,
)
dv.add(f"G5:G{4 + len(PROX_SORTED)}")
ws.add_data_validation(dv)

# Actualizar resumen para incluir las hojas nuevas
ws_res = wb["16_Resumen"]
extra_counts = [
    ("Sello (variantes)",  "=COUNTA('17_Sello'!A:A)-1"),
    ("Proximidad (pares)", "=COUNTA('18_Proximidad_Tematica'!A:A)-1"),
]
start_row = 5 + len(counts)
for i, (label, formula) in enumerate(extra_counts):
    r = start_row + i
    ws_res.cell(row=r, column=1, value=label).font = Font(name="Arial", size=11)
    ws_res.cell(row=r, column=2, value=formula).font = Font(name="Arial", size=11)
    ws_res.cell(row=r, column=1).border = BORDER
    ws_res.cell(row=r, column=2).border = BORDER

wb.save(OUTPUT)
print(f"OK: {OUTPUT}")
print(f"  Líneas: {len(LINEAS)}")
print(f"  Sublíneas: {len(SUBLINEAS)}")
print(f"  Áreas: {len(AREAS)}")
print(f"  Modos: {len(MODOS)}")
print(f"  Salidas: {len(SALIDAS)}")
print(f"  Laboratorios: {len(LABORATORIOS)}")
print(f"  Investigadores: {len(INVESTIGADORES)}")
print(f"  Temas: {len(TEMAS)}")
print(f"  Sublínea↔Tema: {len(SUB_TEMA)}")
print(f"  Lab↔Línea: {len(LAB_LINEA)}")
print(f"  Lab↔Salida: {len(LAB_SALIDA)}")
print(f"  Investigador↔Lab: {len(INV_LAB)}")
print(f"  Investigador↔Modo: {len(INV_MODO)}")
print(f"  Sello variantes: {len(SELLO_VARIANTES)} (1 elegida)")
print(f"  Proximidad pares: {len(PROXIMIDAD_PARES)}")
